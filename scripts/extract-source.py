"""Extract comparable 2020–2025 MHLW job-condition tables."""

from __future__ import annotations

import hashlib
import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

YEARS = list(range(2020, 2026))
SOURCES = {
    "weekend": "https://www.mhlw.go.jp/toukei/list/xls/114-1d-11.xlsx",
    "bonus": "https://www.mhlw.go.jp/toukei/list/xls/114-1d-13.xlsx",
    "commute": "https://www.mhlw.go.jp/toukei/list/xls/114-1d-16.xlsx",
}
INDUSTRIES = [
    ("ALL", "産業計"),
    ("AB", "農林漁業"),
    ("C", "鉱業"),
    ("D", "建設業"),
    ("E", "製造業"),
    ("F", "電気・ガ・熱"),
    ("G", "情報通信"),
    ("H", "運輸業"),
    ("I", "卸売・小売"),
    ("J", "金融・保険"),
    ("K", "不動産"),
    ("L", "学術研究"),
    ("M", "飲食・宿泊"),
    ("N", "生活関連・娯楽"),
    ("O", "教育・学習"),
    ("P", "医療・福祉"),
    ("Q", "複合サービス"),
    ("R", "サービス"),
    ("ST", "公務・その他"),
]


def value(cell: object) -> int:
    if isinstance(cell, bool) or not isinstance(cell, (int, float)):
        raise ValueError(f"expected numeric cell, got {cell!r}")
    number = int(cell)
    if number < 0 or number != cell:
        raise ValueError(f"expected non-negative integer, got {cell!r}")
    return number


def series(sheet: object, row: int, columns: list[int]) -> list[int]:
    return [value(sheet.cell(row, column).value) for column in columns]


def check_industries(sheet: object, column: int) -> None:
    for offset, (industry_id, name) in enumerate(INDUSTRIES):
        source = unicodedata.normalize("NFKC", str(sheet.cell(3 + offset, column).value or "")).strip()
        if industry_id == "ALL":
            if source != name:
                raise ValueError(f"unexpected industry label: {source!r}")
        elif not source.startswith(industry_id) or name not in source:
            raise ValueError(f"unexpected industry label: {source!r}")


def main() -> None:
    if len(sys.argv) != 5:
        raise SystemExit("usage: extract-source.py WEEKEND.xlsx BONUS.xlsx COMMUTE.xlsx OUTPUT_DIR")

    paths = {
        "weekend": Path(sys.argv[1]),
        "bonus": Path(sys.argv[2]),
        "commute": Path(sys.argv[3]),
    }
    output_directory = Path(sys.argv[4])
    books = {
        key: openpyxl.load_workbook(path, data_only=True, read_only=False)
        for key, path in paths.items()
    }

    weekend_old, weekend_current = books["weekend"].worksheets
    bonus_old, bonus_current = books["bonus"].worksheets[2:4]
    commute_old, commute_current = books["commute"].worksheets[2:4]
    for sheet, column in (
        (weekend_current, 1),
        (bonus_current, 2),
        (commute_current, 2),
    ):
        check_industries(sheet, column)

    weekend_bases = [47, 52, 57, 62, 2, 7]
    bonus_columns = [3, 4, 5, 6, 3, 4]
    commute_columns = [3, 4, 5, 6, 3, 4]
    records = []
    for offset, (industry_id, name) in enumerate(INDUSTRIES):
        row = 3 + offset
        weekend_values = {key: [] for key in ("total", "complete", "other", "none")}
        for year_index, base in enumerate(weekend_bases):
            sheet = weekend_old if year_index < 4 else weekend_current
            total, complete, biweekly, other, none = series(sheet, row, list(range(base, base + 5)))
            if total != complete + biweekly + other + none:
                raise ValueError(f"weekend categories do not sum for {industry_id} {YEARS[year_index]}")
            weekend_values["total"].append(total)
            weekend_values["complete"].append(complete)
            weekend_values["other"].append(biweekly + other)
            weekend_values["none"].append(none)

        bonus_values = {"yes": [], "no": []}
        commute_values = {key: [] for key in ("upperLimit", "noLimit", "fixed", "none")}
        for year_index, column in enumerate(bonus_columns):
            bonus_sheet = bonus_old if year_index < 4 else bonus_current
            commute_sheet = commute_old if year_index < 4 else commute_current
            bonus_values["yes"].append(value(bonus_sheet.cell(row, column).value))
            bonus_values["no"].append(value(bonus_sheet.cell(row + 19, column).value))
            for block, key in enumerate(commute_values):
                commute_values[key].append(value(commute_sheet.cell(row + 19 * block, column).value))

        records.append(
            {
                "id": industry_id,
                "name": name,
                "weekend": weekend_values,
                "bonus": bonus_values,
                "commute": commute_values,
            }
        )

    hashes = {key: hashlib.sha256(path.read_bytes()).hexdigest() for key, path in paths.items()}
    index = {
        "asOf": "2026-08-02",
        "edition": "2025年度（令和7年度）",
        "years": YEARS,
        "industryCount": len(INDUSTRIES),
        "recordCount": len(records),
        "employment": "パートタイムを除く常用",
        "sources": [
            {"metric": key, "url": SOURCES[key], "sha256": hashes[key], "bytes": paths[key].stat().st_size}
            for key in ("weekend", "bonus", "commute")
        ],
    }
    output_directory.mkdir(parents=True, exist_ok=True)
    (output_directory / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    (output_directory / "conditions.json").write_text(
        json.dumps(records, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({"hashes": hashes, "industries": len(records), "years": YEARS}, ensure_ascii=False))


if __name__ == "__main__":
    main()
